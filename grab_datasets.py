import itertools
import os
import time
import zipfile

# from rich import print_json
from pathlib import Path, PosixPath

import pandas as pd
import requests
import toml
import xlrd
from bs4 import BeautifulSoup
from loguru import logger
from openpyxl import load_workbook

# Configure loguru logging
logger.remove()
logger.add(
    "logs/grab_datasets.log",
    rotation=100,  # Rotate after 100 log entries
    retention=10,  # Keep 10 files
    level="INFO",
    format="{time:YYYY-MM-DD HH:mm:ss} | {level} | {message}",
)
# Also log to console
logger.add(
    lambda msg: print(msg, end=""),
    level="INFO",
    format="{time:HH:mm:ss} | {level} | {message}",
)

OTHER_VARS_TO_STORE = ["long_name", "code", "short_name", "measure"]
user_agent_header = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
}

# Read local `config.toml` file.
config = toml.load("config.toml")


def get_sheetnames_xlsx(filepath: PosixPath):
    logger.info(f"Loading workbook from {filepath}")
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames


def get_sheetnames_xls(filepath: PosixPath):
    xls = xlrd.open_workbook(filepath, on_demand=True)
    return xls.sheet_names()


def remove_bad_sheets(series: pd.Series):
    return series.apply(lambda x: [el for el in x if "triangle" in el.lower()])


def find_files(url: str):
    try:
        logger.debug(f"Fetching URL: {url}")
        # Add delay to avoid rate limiting
        time.sleep(3)
        response = requests.get(url, headers=user_agent_header)
        logger.debug(f"Response status: {response.status_code}")

        # Check if we got rate limited
        if response.status_code == 429:
            logger.warning(f"Rate limited for {url}, waiting 5 seconds and retrying...")
            time.sleep(5)
            response = requests.get(url, headers=user_agent_header)
            logger.debug(f"Retry response status: {response.status_code}")

        soup = BeautifulSoup(response.text, features="html5lib")

        hrefs = [a["href"] for a in soup.find_all("a")]
        logger.debug(f"Total hrefs found: {len(hrefs)}")

        # Filter files with extensions using pathlib
        valid_extensions = {".xlsx", ".xls", ".zip", ".xlsm"}
        hrefs = [a for a in hrefs if Path(a).suffix.lower() in valid_extensions]
        logger.debug(f"Final filtered files: {len(hrefs)}")
        if hrefs:
            logger.debug(f"First file found: {hrefs[0]}")
        return hrefs
    except Exception as e:
        logger.error(f"Error in find_files for {url}: {e}")
        return []


def download_zip_file(file_url: str, in_file_name: str, short_name: str, code: str):
    """Downloads a zip file from given url.

    :param file_url: url
    :type file_url: str
    :param in_file_name: zip file to download
    :type in_file_name: str
    :return: Name of the file actually downloaded
    :rtype: str
    """
    _ = download_and_save_file(file_url, in_file_name)
    names_to_keep = ["quarterly", "m on m", "1 month", code]
    file_location = Path("scratch") / in_file_name
    try:
        zip_object = zipfile.ZipFile(file_location)
    except zipfile.BadZipFile:
        logger.error(
            f"File {in_file_name} is not a valid zip file, skipping extraction"
        )
        return in_file_name  # Return the original file name
    # Work around introduced because looking for code picks up some cases twice (eg IOS is in both 3M on 3M on M on M)
    names = [name for name in zip_object.namelist() if "3m on 3m" not in name.lower()]
    files_to_extract = [[x for x in names if y in x.lower()] for y in names_to_keep]
    files_to_extract = list(set(itertools.chain(*files_to_extract)))
    # This picks out production or manufacturing which are combined, for some reason,
    # in the Index of Production zip file
    if len(files_to_extract) > 1:
        files_to_extract = [x for x in files_to_extract if short_name in x.lower()]
    for file in files_to_extract:
        zip_object.extract(file, path=Path("scratch"))
    assert len(files_to_extract) == 1
    # Tidy up by removing the zip
    os.remove(file_location)
    return files_to_extract[0]


def download_and_save_file(file_url: str, file_name: str):
    # Download the file from the top of the list
    file_location = Path("scratch") / file_name
    if file_location.is_file():
        logger.info(f"Skipping download of {file_name}; file already exists")
    else:
        # Add delay for downloads too
        time.sleep(3)
        url = "https://www.ons.gov.uk" + file_url
        r = requests.get(url, stream=True, headers=user_agent_header)

        # Check if we got an HTML error page instead of the actual file
        content_type = r.headers.get("content-type", "").lower()
        if "html" in content_type:
            logger.warning(
                f"Got HTML response instead of file for {file_name}, likely rate limited"
            )
            # Wait and try again
            time.sleep(5)
            r = requests.get(url, stream=True, headers=user_agent_header)
            content_type = r.headers.get("content-type", "").lower()
            if "html" in content_type:
                logger.error(f"Still getting HTML response for {file_name}, skipping")
                return None

        with open(Path("scratch") / file_name, "wb") as f:
            f.write(r.content)
    logger.info(f"Success: file download of {file_name} complete")
    return file_name


def convert_yyyy_qn_to_datetime(series: pd.Series):
    return (
        pd.to_datetime(series.apply(lambda x: x[:4] + "-" + str(int(x[-1]) * 3)))
        + pd.offsets.QuarterEnd()
    )


def detect_datetime_format_and_convert(series: pd.Series):
    """
    Intelligently detect datetime format using regex and convert accordingly.
    Handles YYYY QN format, MMM-YY format, and other common formats.
    """
    import re

    if series.empty:
        return series

    # Get first non-null value for format detection
    sample_value = str(series.dropna().iloc[0]).strip()

    # Quarterly format: YYYY QN (e.g., "2023 Q1", "2023Q2")
    if re.match(r"^\d{4}\s*Q[1-4]$", sample_value, re.IGNORECASE):
        return convert_yyyy_qn_to_datetime(series.astype("str").str.strip())

    # Month-Year format: MMM-YY (e.g., "Jan-23", "FEB-24")
    if re.match(r"^[A-Za-z]{3}-\d{2}$", sample_value):
        return pd.to_datetime(series.str.strip(), format="%b-%y", errors="coerce")

    # Try standard datetime conversion as fallback
    return pd.to_datetime(series.astype("str").str.strip(), errors="coerce")


def find_vintage_from_pub_datetime(df_in: pd.DataFrame):
    offsets = {
        "1st": pd.offsets.MonthEnd(),
        "M2": pd.offsets.MonthEnd(2),  # 2nd estimate (month 2)
        "QNA": pd.offsets.QuarterEnd(),
        "M3": pd.offsets.MonthEnd(3),
    }
    df_in["vintage"] = df_in.apply(
        lambda x: offsets[x["estimate"]] + x["pub_datetime"], axis=1
    )
    return df_in


def combined_df_urls(config):
    df_urls = pd.DataFrame()
    frequencies = ["Q", "M"]
    for freq in frequencies:
        df_urls = pd.concat(
            [df_urls, populate_dataframe_of_data_urls(config, freq)], axis=0
        )
    for key, value in config[freq][0].items():
        df_urls[key] = ""
    for freq in frequencies:
        for key, value in config[freq][0].items():
            if key != "urls":
                for inner_key, inner_val in value.items():
                    df_urls.loc[inner_key, key] = inner_val
    return df_urls


def populate_dataframe_of_data_urls(config, freq):
    dict_of_urls = config[freq][0]["urls"]
    dict_of_files = {k: find_files(v) for k, v in dict_of_urls.items()}
    # restrict to only first file found on each page
    for key, value in dict_of_files.items():
        if value:  # Check if list is not empty
            dict_of_files[key] = value[0]
        else:
            logger.warning(f"No files found for {key} at URL: {dict_of_urls[key]}")
            dict_of_files[key] = None
    # Remove entries with None values before creating DataFrame
    dict_of_files_filtered = {k: v for k, v in dict_of_files.items() if v is not None}

    # turn this into a dataframe
    if dict_of_files_filtered:
        df_urls = pd.DataFrame(dict_of_files_filtered, index=["url"]).T
        df_urls["file_name"] = df_urls["url"].apply(lambda x: x.split("/")[-1])
        df_urls[["url", "file_name"]].set_index("url").to_dict()
        df_urls["freq"] = freq
        df_urls["extension"] = df_urls["file_name"].str.split(".").str[1]
    else:
        # Return empty DataFrame with expected columns
        df_urls = pd.DataFrame(columns=["url", "file_name", "freq", "extension"])
    return df_urls


def download_all_files(df_urls):
    # Return early if DataFrame is empty
    if df_urls.empty:
        return df_urls

    # Filter out rows with NaN values in critical columns
    valid_rows = df_urls.dropna(subset=["url", "file_name"])

    if valid_rows.empty:
        logger.warning("No valid rows to download after filtering NaN values")
        return df_urls

    df_urls["dl_filename"] = ""
    # Download non-zips
    query = valid_rows["extension"] != "zip"
    if query.any():
        df_urls.loc[valid_rows[query].index, "dl_filename"] = valid_rows.loc[
            query, :
        ].apply(lambda x: download_and_save_file(x["url"], x["file_name"]), axis=1)
    # Download zips
    zip_query = ~query
    if zip_query.any():
        df_urls.loc[valid_rows[zip_query].index, "dl_filename"] = valid_rows.loc[
            zip_query, :
        ].apply(
            lambda x: download_zip_file(
                x["url"], x["file_name"], x["short_name"], x["code"]
            ),
            axis=1,
        )
    df_urls["dl_fn_extension"] = df_urls["dl_filename"].str.split(".").str[1]
    return df_urls


def nominate_sheets_from_ss(df_urls):
    # Add sheet names
    df_urls["sheet_names"] = "None"
    df_urls.loc[df_urls["dl_fn_extension"] == "xlsx", "sheet_names"] = df_urls.loc[
        df_urls["dl_fn_extension"] == "xlsx", :
    ].apply(lambda x: get_sheetnames_xlsx(Path("scratch") / x["dl_filename"]), axis=1)
    if "xlsm" in df_urls["dl_fn_extension"].unique():
        df_urls.loc[df_urls["dl_fn_extension"] == "xlsm", "sheet_names"] = df_urls.loc[
            df_urls["dl_fn_extension"] == "xlsm", :
        ].apply(
            lambda x: get_sheetnames_xlsx(Path("scratch") / x["dl_filename"]), axis=1
        )
    if "xls" in df_urls["dl_fn_extension"].unique():
        df_urls.loc[df_urls["dl_fn_extension"] == "xls", "sheet_names"] = df_urls.loc[
            df_urls["dl_fn_extension"] == "xls", :
        ].apply(
            lambda x: get_sheetnames_xls(Path("scratch") / x["dl_filename"]), axis=1
        )
    df_urls["sheet_names"] = remove_bad_sheets(df_urls["sheet_names"])
    # stick only to the first sheet
    df_urls["sheet_names"] = df_urls["sheet_names"].apply(lambda x: x[0])
    return df_urls


def enforce_types(df):
    # Ensure the correct types are enforced
    type_dict = {
        "long_name": "category",
        "code": "category",
        "short_name": "category",
        "measure": "category",
    }
    for key, value in type_dict.items():
        df[key] = df[key].astype(value)
    return df


def process_triangle_file(df_urls_row):
    logger.info(f"Processing {df_urls_row.name}")
    file_name, sheet_name = df_urls_row["dl_filename"], df_urls_row["sheet_names"]
    df = pd.read_excel(Path("scratch") / file_name, sheet_name=sheet_name)
    # Remove all the of the guff
    search_text = "Relating to Period"
    alt_search_text = search_text + " (three months ending)"
    alt_alt_search_text = "Relating to period"
    df = df.dropna(how="all", axis=1).dropna(how="all", axis=0)
    # work around for variations on 'relating to period'
    dates_row = (
        df[(df == search_text) | (df == alt_search_text) | (df == alt_alt_search_text)]
        .dropna(how="all", axis=1)
        .dropna(how="all", axis=0)
        .index.values
    )
    df = df.rename(columns=dict(zip(df.columns, df.loc[dates_row, :].values[0])))
    # remove any lingering first cols
    if search_text in list(df.columns):
        srch_txt_ix = list(df.columns).index(search_text)
    elif alt_search_text in list(df.columns):
        srch_txt_ix = list(df.columns).index(alt_search_text)
        df = df.rename(columns={df.columns[srch_txt_ix]: search_text})
    elif alt_alt_search_text in list(df.columns):
        srch_txt_ix = list(df.columns).index(alt_alt_search_text)
        df = df.rename(columns={df.columns[srch_txt_ix]: search_text})
    else:
        raise ValueError(
            "None of the names associated with dates can be found in the spreadsheet"
        )
    if srch_txt_ix != 0:
        df = df[df.columns[srch_txt_ix:]].copy()
    format_datetime = "%Y-%m-%d"
    if any([x in df_urls_row["code"] for x in ["abjr", "npqt", "ihyq", "exp", "imp"]]):
        format_datetime = "%b-%y"
    df[df.columns[0]] = pd.to_datetime(
        df[df.columns[0]], errors="coerce", format=format_datetime
    )
    first_datetime_row = (
        pd.to_datetime(df[df.columns[0]], errors="coerce", format=format_datetime)
        .dropna()
        .index.min()
    )
    df = df.loc[first_datetime_row:, :]
    # fill in the "latest estimate" entry with a datetime
    df = df[~pd.isna(df[search_text])].copy()
    time_series_down = pd.to_datetime(df[search_text], errors="coerce")
    time_series_down.iloc[-1] = time_series_down.iloc[-2] + pd.DateOffset(months=3)
    df[search_text] = time_series_down
    df = pd.melt(df, id_vars=search_text, var_name="datetime")
    df = df.rename(columns={search_text: "vintage"})
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    # Intelligently detect and convert datetime format using regex
    df["datetime"] = detect_datetime_format_and_convert(df["datetime"])
    df = df.dropna(subset=["value"])
    for var in OTHER_VARS_TO_STORE:
        df[var] = df_urls_row[var]
    return enforce_types(df)


def get_ons_series(code: str):
    # First get the uri for the time series using the search API
    search_url = (
        f"https://api.beta.ons.gov.uk/v1/search?content_type=timeseries&cdids={code}"
    )
    search_response = requests.get(search_url).json()

    # Extract the uri from the search results
    if "items" not in search_response or len(search_response["items"]) == 0:
        raise ValueError(f"Time series {code} not found")

    uri = search_response["items"][0]["uri"]

    # Build the data URL using the new API structure
    url = f"https://api.beta.ons.gov.uk/v1/data?uri={uri}"

    # Get the data from the ONS API:
    json_data = requests.get(url).json()

    # Prep the data for a quick plot
    title = json_data["description"]["title"]
    df = (
        pd.DataFrame(pd.json_normalize(json_data["months"]))
        .assign(
            date=lambda x: pd.to_datetime(x["date"], format="%Y %b"),
            value=lambda x: pd.to_numeric(x["value"]),
        )
        .set_index("date")
    )
    df["title"] = title
    return df


def populate_nonrev_series(series_name: str):
    xf = get_ons_series(
        config["nonrev"][0]["code"][series_name],
    )
    xf = xf.reset_index()
    xf["vintage"] = xf["date"]
    for var in OTHER_VARS_TO_STORE:
        xf[var] = config["nonrev"][0][var][series_name]
    xf = xf.drop(
        ["label", "month", "quarter", "sourceDataset", "updateDate", "year", "title"],
        axis=1,
    )
    xf = xf.rename(columns={"date": "datetime"})
    return xf


def get_all_non_rev_series(config):
    xf = pd.DataFrame()
    for name in config["nonrev"][0]["dataset"].keys():
        temp_df = populate_nonrev_series(name)
        xf = pd.concat([xf, temp_df], axis=0)
    return enforce_types(xf)


# Get the urls of the revisions & downloaded them
df_urls = combined_df_urls(config)
df_urls = download_all_files(df_urls)
df_urls = nominate_sheets_from_ss(df_urls)

# Extract the data from the files and combine
df = pd.concat(
    [process_triangle_file(df_urls.iloc[i]) for i in range(len(df_urls))], axis=0
)

# Pick up the non-revised data
xf = get_all_non_rev_series(config)
df = pd.concat([df, xf], axis=0)


# Prep to write to file
cat_cols = ["long_name", "code", "short_name", "measure"]
for col in cat_cols:
    df[col] = df[col].astype("category")

# save to file
df.to_parquet(Path("scratch/realtimedata.parquet"))
